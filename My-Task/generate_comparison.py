import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── style helpers ─────────────────────────────────────────────────────────────
C_HEADER  = "1F4E79"
C_SUB     = "2E75B6"
C_DIFF    = "FFD966"
C_MISSING = "FF7043"
C_OK      = "E2EFDA"
C_RED     = "FFCCCC"
C_WHITE   = "FFFFFF"
C_GRAY    = "F2F2F2"

def fill(c): return PatternFill("solid", fgColor=c)
def bd():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def hcell(ws, r, c, v, bg=C_HEADER, fg="FFFFFF", sz=11, wrap=False, bold=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = bd()
    return cell

def dcell(ws, r, c, v, bg=C_WHITE, bold=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=10)
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = bd()
    return cell

def autowidth(ws, mn=10, mx=45):
    for col in ws.columns:
        w = mn
        for cell in col:
            if cell.value:
                w = max(w, min(len(str(cell.value)) + 2, mx))
        ws.column_dimensions[get_column_letter(col[0].column)].width = w

# ══════════════════════════════════════════════════════════════════════════════
# DATA
# ══════════════════════════════════════════════════════════════════════════════

FILES = ["adm", "d-inspect", "d-execute", "d-order"]   # adm = master/reference

summary_rows = [
    # metric                    adm(master) d-inspect   d-execute   d-order     note
    ("Total Records",           283,        283,        283,        278,        "d-order kurang 5 record dari master ADM"),
    ("Timestamp Precision",     "datetime(3)", "datetime2(7)", "datetime2(7)", "datetime2(7)",
                                                                "adm (master) menggunakan presisi 3 desimal, slave 7 desimal"),
    ("Records Missing vs Master",0,         0,          0,          5,          "d-order tidak memiliki 5 record terbaru dari ADM"),
    ("is_active Mismatch vs ADM",0,         0,          0,          "3+",       "d-order memiliki nilai is_active berbeda dari master"),
    ("valid_from Mismatch vs ADM",0,        0,          0,          "3+",       "d-order memiliki valid_from berbeda dari master"),
    ("valid_to Mismatch vs ADM", 0,         0,          0,          "3+",       "d-order memiliki valid_to berbeda dari master"),
    ("employee_name Mismatch",  "—",        1,          0,          0,          "d-inspect row 163 berbeda dari master ADM (nama lebih panjang)"),
    ("Sync Status",             "MASTER",   "✓ Sync",   "✓ Sync",   "✗ NOT SYNC",
                                                                "d-order tidak tersinkronisasi dengan master ADM"),
]

# Records missing in d-order (exist in d-inspect/adm/d-execute)
missing_rows = [
    # id  pno        name                    section  section_desc           valid_from    valid_to       is_active  note
    (215, 10030993, "Abdul Gaffar",          "S096",  "Supp. Gear Equipment","2025-11-17", "2049-12-18",  1,  "Ada di d-inspect/adm/d-execute, tidak ada di d-order"),
    (247, 10012653, "Abdul Ghofar",          "S096",  "Supp. Gear Equipment","2026-06-02", "2026-06-03",  0,  "Ada di d-inspect/adm/d-execute, tidak ada di d-order"),
    (248, 10007517, "Darmawan",              "S096",  "Supp. Gear Equipment","2026-05-01", "2049-05-01",  1,  "Ada di d-inspect/adm/d-execute, tidak ada di d-order"),
    (249, 10028881, "Andi Winarko",          "S001",  "Non-Section",         "2026-05-19", "2042-12-31",  1,  "Ada di d-inspect/adm/d-execute, tidak ada di d-order"),
    (283, 10012653, "Abdul Ghofar",          "S096",  "Supp. Gear Equipment","2026-06-02", "2050-12-31",  1,  "Record terbaru (hari ini), tidak ada di d-order"),
]

# Business data differences (ref = d-inspect/adm/d-execute)
# id, pno, name, section, field, val_inspect, val_adm, val_execute, val_order
biz_diff = [
    # employee_name diff in adm
    (163, 10005004, "Arif Himawanto Novan Saputra / Arif Himawanto", "S095",
     "employee_name",
     "Arif Himawanto Novan Saputra", "Arif Himawanto", "Arif Himawanto Novan Saputra", "Arif Himawanto Novan Saputra"),

    # is_active diffs in d-order
    (214, 10030993, "Abdul Gaffar", "S103",
     "is_active", "0", "0", "0", "1"),

    (220, 10011508, "Fauzy Kusuma Nur Handy", "S096",
     "is_active", "0", "0", "0", "1"),

    (221, 10011508, "Fauzy Kusuma Nur Handy", "S094",
     "is_active", "0", "0", "0", "1"),

    (254, 10008309, "Wagiyo", "S094",
     "is_active", "0", "0", "0", "1"),

    # valid_from diffs in d-order
    (242, 10028885, "Muhammad Nor", "S001",
     "valid_from", "2026-05-19", "2026-05-19", "2026-05-19", "2026-04-10"),

    (242, 10028885, "Muhammad Nor", "S001",
     "valid_to", "2044-12-31", "2044-12-31", "2044-12-31", "2044-04-10"),

    (246, 10028881, "Andi Winarko", "S001",
     "valid_from", "2026-05-19", "2026-05-19", "2026-05-19", "2026-05-11"),

    (246, 10028881, "Andi Winarko", "S001",
     "valid_to", "2042-12-31", "2042-12-31", "2042-12-31", "2040-05-11"),

    (247, 10004843, "Gusti Yanuar Ari Rivani", "S001",
     "valid_from", "2026-05-19", "2026-05-19", "2026-05-19", "2026-05-11"),

    (247, 10004843, "Gusti Yanuar Ari Rivani", "S001",
     "valid_to", "2040-12-31", "2040-12-31", "2040-12-31", "2040-05-11"),

    (254, 10008309, "Wagiyo", "S094",
     "valid_from", "2026-05-19", "2026-05-19", "2026-05-19", "2026-05-18"),
]

# d-order structure differences (different md_config_section_id mapping)
structure_diff = [
    # (id_others, id_dorder, pno, name, section, note)
    (215, "N/A",  10030993, "Abdul Gaffar",       "S096", "Record ini ada di others ID 215, tidak ada di d-order sama sekali"),
    (221, 220,    10011508, "Fauzy Kusuma",        "S094", "ID berbeda: others=222 (Fauzy S094 Jan), d-order=221"),
    (222, 221,    10009744, "Tudi",                "S099", "ID shift: others=224, d-order=222"),
    (223, "N/A",  10011508, "Fauzy Kusuma",        "S094", "Record Fauzy S094 May-21 (others ID 223) tidak ada di d-order"),
    (247, "N/A",  10012653, "Abdul Ghofar",        "S096", "Record Abdul Ghofar S096 Jun-02 is_active=0 tidak ada di d-order"),
    (283, "N/A",  10012653, "Abdul Ghofar",        "S096", "Record terbaru (hari ini) tidak ada di d-order"),
]

# ══════════════════════════════════════════════════════════════════════════════
# BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── SHEET 1: SUMMARY ──────────────────────────────────────────────────────────
ws = wb.create_sheet("1. Summary")
ws.freeze_panes = "A4"
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 14
ws.row_dimensions[3].height = 22

ws.merge_cells("A1:G1")
c = ws["A1"]
c.value = "DB Comparison Report — md_config_section   |   Master: ADM"
c.fill = fill(C_HEADER)
c.font = Font(bold=True, color="FFFFFF", size=16)
c.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A2:G2")
c = ws["A2"]
c.value = "Master DB: ADM   |   Slave: d-inspect | d-execute | d-order   |   Generated: 2026-06-02   |   Excluded: created_on, changed_on"
c.fill = fill("D6E4F7")
c.font = Font(italic=True, size=9, color="333333")
c.alignment = Alignment(horizontal="center")

headers = ["Metric", "adm (MASTER)", "d-inspect", "d-execute", "d-order", "Status", "Note"]
for ci, h in enumerate(headers, 1):
    hcell(ws, 3, ci, h, bg=C_SUB)

for ri, row in enumerate(summary_rows, 4):
    ws.row_dimensions[ri].height = 18
    metric, vi, va, ve, vo, note = row
    bg = C_GRAY if ri % 2 == 0 else C_WHITE
    dcell(ws, ri, 1, metric, bg, bold=True)
    for ci, v in enumerate([vi, va, ve, vo], 2):
        cell_bg = bg
        if metric in ("Total Records","Records Missing vs Ref","is_active Mismatch",
                      "valid_from Mismatch","valid_to Mismatch","employee_name Mismatch"):
            if v not in (0, "0", 283) and ci == 5:
                cell_bg = C_DIFF
        if metric == "Sync Status" and "OUT" in str(v):
            cell_bg = C_MISSING
        dcell(ws, ri, ci, v, cell_bg)
    # status col
    if metric == "Sync Status":
        statuses = [("✓ OK", C_OK), ("✓ OK", C_OK), ("✓ OK", C_OK), ("✗ NOT SYNC", C_MISSING)]
    else:
        statuses = [(None, bg)] * 4
    if metric == "Sync Status":
        for ci2, (sv, sbg) in enumerate(statuses, 2):
            dcell(ws, ri, ci2, sv, sbg)
    dcell(ws, ri, 6, "⚠ Issue" if any(str(v) not in ("0","✓ OK") for v in [vi,va,ve,vo] if v != vi) else "✓ OK",
          C_DIFF if any(str(v) not in ("0","✓ OK") for v in [vi,va,ve,vo] if str(v) != str(vi)) else C_OK)
    dcell(ws, ri, 7, note, bg)

autowidth(ws)

# ── SHEET 2: MISSING IN d-order ───────────────────────────────────────────────
ws2 = wb.create_sheet("2. Missing in d-order")
ws2.freeze_panes = "A4"
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[3].height = 22

ws2.merge_cells("A1:J1")
c = ws2["A1"]
c.value = "Records yang ADA di d-inspect / adm / d-execute tapi TIDAK ADA di d-order"
c.fill = fill(C_MISSING)
c.font = Font(bold=True, color="FFFFFF", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")

ws2.merge_cells("A2:J2")
c = ws2["A2"]
c.value = f"Total: {len(missing_rows)} record missing"
c.fill = fill("FFE0D0")
c.font = Font(italic=True, size=10)
c.alignment = Alignment(horizontal="center")

h2 = ["ref_id","personnel_number","employee_name","section","section_desc",
      "valid_from","valid_to","is_active","d-inspect","adm","d-execute","d-order","Note"]
for ci, h in enumerate(h2, 1):
    hcell(ws2, 3, ci, h, bg=C_SUB)

for ri, row in enumerate(missing_rows, 4):
    ws2.row_dimensions[ri].height = 16
    rid, pno, name, sec, desc, vf, vt, ia, note = row
    vals = [rid, pno, name, sec, desc, vf, vt, ia, "✓", "✓", "✓", "✗ MISSING"]
    for ci, v in enumerate(vals, 1):
        bg = C_MISSING if ci == 12 else (C_OK if ci in (9,10,11) else C_WHITE)
        dcell(ws2, ri, ci, v, bg)
    dcell(ws2, ri, 13, note, "FFF3E0")

autowidth(ws2)

# ── SHEET 3: BUSINESS DATA DIFFERENCES ───────────────────────────────────────
ws3 = wb.create_sheet("3. Business Data Diff")
ws3.freeze_panes = "A4"
ws3.row_dimensions[1].height = 30
ws3.row_dimensions[3].height = 22

ws3.merge_cells("A1:I1")
c = ws3["A1"]
c.value = "Business Data Differences (is_active / valid_from / valid_to / employee_name)"
c.fill = fill(C_DIFF.replace("FF","CC"))
c.font = Font(bold=True, size=13)
c.alignment = Alignment(horizontal="center", vertical="center")

ws3.merge_cells("A2:I2")
c = ws3["A2"]
c.value = f"Total perbedaan: {len(biz_diff)} field   |   🟡 = berbeda dari mayoritas   |   🟢 = sesuai mayoritas"
c.fill = fill("FFFDE7")
c.font = Font(italic=True, size=10)
c.alignment = Alignment(horizontal="center")

h3 = ["id","personnel_number","employee_name","section","field",
      "d-inspect","adm","d-execute","d-order"]
for ci, h in enumerate(h3, 1):
    hcell(ws3, 3, ci, h, bg=C_SUB)

for ri, row in enumerate(biz_diff, 4):
    ws3.row_dimensions[ri].height = 16
    rid, pno, name, sec, field, vi, va, ve, vo = row
    vals_data = [rid, pno, name, sec, field]
    file_vals = [vi, va, ve, vo]
    from collections import Counter
    majority = Counter(file_vals).most_common(1)[0][0]
    for ci, v in enumerate(vals_data, 1):
        dcell(ws3, ri, ci, v, C_GRAY if ri % 2 == 0 else C_WHITE)
    for ci2, v in enumerate(file_vals, 6):
        bg = C_DIFF if v != majority else C_OK
        dcell(ws3, ri, ci2, v, bg)

autowidth(ws3)

# ── SHEET 4: STRUCTURE DIFFERENCES ───────────────────────────────────────────
ws4 = wb.create_sheet("4. Structure Diff d-order")
ws4.freeze_panes = "A4"
ws4.row_dimensions[1].height = 30
ws4.row_dimensions[3].height = 22

ws4.merge_cells("A1:G1")
c = ws4["A1"]
c.value = "Perbedaan Struktur / ID Mapping di d-order"
c.fill = fill(C_HEADER)
c.font = Font(bold=True, color="FFFFFF", size=13)
c.alignment = Alignment(horizontal="center", vertical="center")

ws4.merge_cells("A2:G2")
c = ws4["A2"]
c.value = "d-order memiliki ID yang berbeda karena beberapa record tidak ada, menyebabkan pergeseran urutan"
c.fill = fill("D6E4F7")
c.font = Font(italic=True, size=10)
c.alignment = Alignment(horizontal="center")

h4 = ["id (d-inspect/adm/d-execute)","id (d-order)","personnel_number",
      "employee_name","section","Note"]
for ci, h in enumerate(h4, 1):
    hcell(ws4, 3, ci, h, bg=C_SUB)

for ri, row in enumerate(structure_diff, 4):
    ws4.row_dimensions[ri].height = 16
    id_others, id_dorder, pno, name, sec, note = row
    vals = [id_others, id_dorder, pno, name, sec, note]
    for ci, v in enumerate(vals, 1):
        bg = C_MISSING if str(v) == "N/A" else (C_GRAY if ri % 2 == 0 else C_WHITE)
        dcell(ws4, ri, ci, v, bg)

autowidth(ws4)

# ── SHEET 5: CONCLUSION ───────────────────────────────────────────────────────
ws5 = wb.create_sheet("5. Conclusion & Action")
ws5.column_dimensions["A"].width = 30
ws5.column_dimensions["B"].width = 80

ws5.row_dimensions[1].height = 36
ws5.merge_cells("A1:B1")
c = ws5["A1"]
c.value = "Conclusion & Recommended Actions"
c.fill = fill(C_HEADER)
c.font = Font(bold=True, color="FFFFFF", size=14)
c.alignment = Alignment(horizontal="center", vertical="center")

conclusions = [
    ("MECHANISM",  "API mentrigger save ke semua DB sekaligus (ADM, d-inspect, d-execute, d-order) — bukan sync dari ADM. ADM adalah tabel utama/master."),
    ("FINDING",    "ADM, d-inspect, d-execute → Data praktis identik. Slave DB berhasil menerima trigger dari API dengan benar."),
    ("FINDING",    "d-order → TIDAK TERUPDATE: 5 record missing, 4+ nilai is_active salah, 4+ nilai valid_from/valid_to berbeda dari ADM."),
    ("ROOT CAUSE", "API tidak mentrigger save ke d-order, atau trigger ke d-order gagal secara silent (tidak ada error handling yang tampil ke user)."),
    ("ROOT CAUSE", "Validasi tanggal di SP terlalu ketat untuk UPDATE: 'IF @validFrom < @today' menyebabkan rollback saat edit record lama, sehingga d-order tidak terupdate."),
    ("ROOT CAUSE", "Kemungkinan d-order menggunakan koneksi DB atau SP yang berbeda dari 3 DB lainnya, dan tidak dipanggil oleh API."),
    ("ACTION 1",   "Cek API layer: pastikan trigger save ke d-order dipanggil dan tidak di-skip secara kondisional."),
    ("ACTION 2",   "Fix SP usp_post_adm_config_section: validasi tanggal masa lalu hanya untuk INSERT (configSectionId = 0), bukan UPDATE."),
    ("ACTION 3",   "Tambahkan error handling eksplisit di API untuk trigger ke d-order agar error tidak silent."),
    ("ACTION 4",   "Patch data d-order: update 5 record missing dan perbaiki is_active / valid_from / valid_to yang salah."),
    ("ACTION 5",   "Monitoring: tambahkan cek otomatis jumlah record antar DB setelah setiap operasi save dari API."),
    ("NOTE",       "d-inspect row 163: employee_name berbeda dari ADM master ('Arif Himawanto Novan Saputra' vs 'Arif Himawanto') — kemungkinan update manual langsung ke d-inspect."),
]

colors = {"MECHANISM": "D6E4F7", "FINDING": "D6E4F7", "ROOT CAUSE": C_DIFF,
          "ACTION 1": C_OK, "ACTION 2": C_OK, "ACTION 3": C_OK,
          "ACTION 4": C_OK, "ACTION 5": C_OK, "NOTE": "FFF9C4"}

for ri, (tag, text) in enumerate(conclusions, 3):
    ws5.row_dimensions[ri].height = 30
    bg = colors.get(tag, C_WHITE)
    dcell(ws5, ri, 1, tag, bg, bold=True)
    c2 = ws5.cell(row=ri, column=2, value=text)
    c2.fill = fill(bg)
    c2.font = Font(size=10)
    c2.alignment = Alignment(vertical="center", wrap_text=True)
    c2.border = bd()

# save
out = r"d:\OneDrive-Btech\OneDrive - Bukit Technology\Notebooks\office-notes\My-Task\md_config_section_comparison.xlsx"
wb.save(out)
print(f"Done! Saved: {out}")
